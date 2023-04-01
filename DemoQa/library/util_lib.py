import os

import openpyxl

from DemoQa.library import config
from DemoQa.library.config import excel_file_path


def create_result_directory():
    result_dir = config.RESULT_PATH
    if not os.path.exists(result_dir):
        os.mkdir(result_dir)
    return result_dir


def create_screenshot_directory(result_dir):
    screenshot_path = result_dir + "/screenshot"
    if not os.path.exists(screenshot_path):
        os.mkdir(screenshot_path)
    return screenshot_path


def create_html_report(html_report_path):
    f = open(html_report_path + "/"'HtmlReport.html', 'w')
    f.close()


def create_sheet(sheet=None):
    workbook = openpyxl.load_workbook(excel_file_path)
    if not sheet:
        sheet = workbook.create_sheet("arjun")
        print(sheet)
        print("arjun created")
    else:
        sheet = workbook.get_sheet_by_name("arjun")
        print("arjun sheet already exists")


dir = create_result_directory()
print(dir)

create_sheet()