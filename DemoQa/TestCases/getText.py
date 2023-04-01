import time
from copy import copy

import openpyxl
import xlrd
import xlwt
from selenium.webdriver.common.by import By
from DemoQa.library.config import driver, excel_file_path


def get_text_after_test(sheet_name, locator, split_by_text):
    print("function STARTED : get_text_after_test ")
    time.sleep(1)
    try:
        text_element = driver.find_element(By.XPATH, locator).text
        print(text_element)
        text_element = text_element.split(split_by_text)
        text_element_after_split_by_text = text_element[1]
        text_element_after_space_removed = text_element_after_split_by_text.strip()
        print("Total Count of Computers  : " + text_element_after_space_removed)

        # open workbook
        workbook = openpyxl.load_workbook(excel_file_path)

        if "ComputerCount" in workbook.sheetnames:
            print("ComputerCount sheet already exists")
            sheet = workbook.get_sheet_by_name("ComputerCount")
            sheet["A1"] = "Total Computers count :"
            sheet["B1"] = text_element_after_space_removed
        else:
            # create the sheet
            sheet = workbook.create_sheet("ComputerCount")
            sheet["A1"] = "Total Computers count :"
            sheet["B1"] = text_element_after_space_removed

        # save the workbook
        workbook.save(excel_file_path)

    except Exception as e:
        print(e)
